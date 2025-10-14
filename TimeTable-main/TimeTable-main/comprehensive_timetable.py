import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import csv
import glob
import os
import json

def load_config():
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return config['duration_constants']
    except:
        return {
            'hour_slots': 2,
            'lecture_duration': 3,
            'lab_duration': 4,
            'tutorial_duration': 2,
            'self_study_duration': 2, 
            'break_duration': 1
        }

durations = load_config()
HOUR_SLOTS = durations['hour_slots']
LECTURE_DURATION = durations['lecture_duration']
LAB_DURATION = durations['lab_duration']
TUTORIAL_DURATION = durations['tutorial_duration']
SELF_STUDY_DURATION = durations['self_study_duration']
BREAK_DURATION = durations['break_duration']

DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)

LUNCH_WINDOW_START = time(12, 30)
LUNCH_WINDOW_END = time(14, 0)
LUNCH_DURATION = 60

TIME_SLOTS = []
lunch_breaks = {}

def calculate_lunch_breaks(semesters):
    global lunch_breaks
    lunch_breaks = {}
    total_semesters = len(semesters)
    if total_semesters == 0:
        return lunch_breaks
    total_window_minutes = (
        LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
        LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute
    )
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    sorted_semesters = sorted(semesters)
    for i, semester in enumerate(sorted_semesters):
        start_minutes = (LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + int(i * stagger_interval))
        start_hour = start_minutes // 60
        start_min = start_minutes % 60
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour = end_minutes // 60
        end_min = end_minutes % 60
        lunch_breaks[semester] = (
            time(start_hour, start_min),
            time(end_hour, end_min)
        )
    return lunch_breaks

def initialize_time_slots():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        slots.append((current, next_time.time()))
        current_time = next_time
    return slots

def load_rooms():
    rooms = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except FileNotFoundError:
        print("Warning: rooms.csv not found, using default room allocation")
        return None
    return rooms

def load_batch_data():
    batch_info = {}
    try:
        df = pd.read_csv('combined.csv')
        grouped = df.groupby(['Department', 'Semester'])
        for (dept, sem), group in grouped:
            if 'total_students' in group.columns and not group['total_students'].isna().all():
                total_students = int(group['total_students'].max())
                max_batch_size = 70
                num_sections = (total_students + max_batch_size - 1) // max_batch_size
                section_size = (total_students + num_sections - 1) // num_sections
                batch_info[(dept, sem)] = {
                    'total': total_students,
                    'num_sections': num_sections,
                    'section_size': section_size
                }
        basket_courses = df[df['Course Code'].astype(str).str.contains('^B[0-9]')]
        for _, course in basket_courses.iterrows():
            code = str(course['Course Code'])
            if 'total_students' in df.columns and pd.notna(course['total_students']):
                total_students = int(course['total_students'])
            else:
                total_students = 35
            batch_info[('ELECTIVE', code)] = {
                'total': total_students,
                'num_sections': 1,
                'section_size': total_students
            }
    except FileNotFoundError:
        print("Warning: combined.csv not found, using default batch sizes")
    except Exception as e:
        print(f"Warning: Error processing batch sizes from combined.csv: {e}")
    return batch_info

def find_adjacent_lab_room(room_id, rooms):
    if not room_id:
        return None
    current_num = int(''.join(filter(str.isdigit, rooms[room_id]['roomNumber'])))
    current_floor = current_num // 100
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num = int(''.join(filter(str.isdigit, room['roomNumber'])))
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, course_code="", used_rooms=None):
    if not rooms:
        return "DEFAULT_ROOM"
    required_capacity = 60
    is_basket = is_basket_course(course_code)
    total_students = None
    try:
        df = pd.read_csv('combined.csv')
        if course_code and not is_basket:
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                try:
                    val = course_row['total_students'].iloc[0]
                    if pd.notna(val) and str(val).isdigit():
                        total_students = int(val)
                except (ValueError, TypeError):
                    pass
        elif is_basket:
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                try:
                    val = course_row['total_students'].iloc[0]
                    if pd.notna(val) and str(val).isdigit():
                        total_students = int(val)
                    else:
                        elective_info = batch_info.get(('ELECTIVE', course_code))
                        if elective_info:
                            total_students = elective_info['section_size']
                except (ValueError, TypeError):
                    elective_info = batch_info.get(('ELECTIVE', course_code))
                    if elective_info:
                        total_students = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                total_students = dept_info['section_size']
    except Exception as e:
        print(f"Warning: Error getting total_students from combined.csv: {e}")
    if total_students:
        required_capacity = total_students
    elif batch_info:
        if is_basket:
            elective_info = batch_info.get(('ELECTIVE', course_code))
            if elective_info:
                required_capacity = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                required_capacity = dept_info['section_size']
    used_room_ids = set() if used_rooms is None else used_rooms
    if course_type in ['LEC', 'TUT', 'SS'] and required_capacity > 70:
        seater_120_rooms = {rid: room for rid, room in rooms.items() if 'SEATER_120' in room['type'].upper()}
        if required_capacity > 120:
            seater_240_rooms = {rid: room for rid, room in rooms.items() if 'SEATER_240' in room['type'].upper()}
            room_id = try_room_allocation(seater_240_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
            if room_id:
                return room_id
        room_id = try_room_allocation(seater_120_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
        if room_id:
            return room_id
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        dept_info = batch_info.get((department, semester))
        if dept_info and dept_info['total'] > 35:
            for room_id, room in rooms.items():
                if room_id in used_room_ids or room['type'].upper() != course_type:
                    continue
                slots_free = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        slots_free = False
                        break
                if slots_free:
                    adjacent_room = find_adjacent_lab_room(room_id, rooms)
                    if adjacent_room and adjacent_room not in used_room_ids:
                        adjacent_free = True
                        for i in range(duration):
                            if start_slot + i in rooms[adjacent_room]['schedule'][day]:
                                adjacent_free = False
                                break
                        if adjacent_free:
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                                rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                            return f"{room_id},{adjacent_room}"
        return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)
    if course_type in ['LEC', 'TUT', 'SS'] or is_basket:
        lecture_rooms = {rid: room for rid, room in rooms.items() if 'LECTURE_ROOM' in room['type'].upper()}
        seater_rooms = {rid: room for rid, room in rooms.items() if 'SEATER' in room['type'].upper()}
        if is_basket:
            basket_group = get_basket_group(course_code)
            basket_used_rooms = set()
            basket_group_rooms = {}
            room_usage = {rid: sum(len(room['schedule'][d]) for d in range(len(DAYS))) for rid, room in rooms.items()}
            sorted_lecture_rooms = dict(sorted(lecture_rooms.items(), key=lambda x: room_usage[x[0]]))
            sorted_seater_rooms = dict(sorted(seater_rooms.items(), key=lambda x: room_usage[x[0]]))
            for room_dict in [sorted_lecture_rooms, sorted_seater_rooms]:
                for room_id, room in room_dict.items():
                    is_used = False
                    for slot in range(start_slot, start_slot + duration):
                        if slot in rooms[room_id]['schedule'][day]:
                            if slot in timetable[day]:
                                slot_data = timetable[day][slot]
                                if (slot_data['classroom'] == room_id and slot_data['type'] is not None):
                                    slot_code = slot_data.get('code', '')
                                    if get_basket_group(slot_code) == basket_group:
                                        basket_group_rooms[slot_code] = room_id
                                    else:
                                        basket_used_rooms.add(room_id)
                            is_used = True
                            break
                    if not is_used and room_id not in basket_used_rooms:
                        if 'capacity' in room and room['capacity'] >= required_capacity:
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                            return room_id
            if course_code in basket_group_rooms:
                return basket_group_rooms[course_code]
            room_id = try_room_allocation(lecture_rooms, 'LEC', required_capacity, day, start_slot, duration, basket_used_rooms)
            if not room_id:
                room_id = try_room_allocation(seater_rooms, 'LEC', required_capacity, day, start_slot, duration, basket_used_rooms)
            if room_id:
                basket_group_rooms[course_code] = room_id
            return room_id
        room_id = try_room_allocation(lecture_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
        if not room_id:
            room_id = try_room_allocation(seater_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
        return room_id
    return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)

def try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids):
    for room_id, room in rooms.items():
        if room_id in used_room_ids or room['type'].upper() == 'LIBRARY':
            continue
        if course_type in ['LEC', 'TUT', 'SS']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
        elif course_type == 'COMPUTER_LAB' and room['type'].upper() != 'COMPUTER_LAB':
            continue
        elif course_type == 'HARDWARE_LAB' and room['type'].upper() != 'HARDWARE_LAB':
            continue
        if course_type not in ['COMPUTER_LAB', 'HARDWARE_LAB'] and room['capacity'] < required_capacity:
            continue
        slots_free = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                slots_free = False
                break
        if slots_free:
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
    return None

def get_required_room_type(course):
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'
    else:
        return 'LECTURE_ROOM'

def is_basket_course(code):
    return code.startswith('B') and '-' in code

def get_basket_group(code):
    if is_basket_course(code):
        return code.split('-')[0]
    return None

def get_basket_group_slots(timetable, day, basket_group):
    basket_slots = []
    for slot_idx, slot in timetable[day].items():
        code = slot.get('code', '')
        if code and get_basket_group(code) == basket_group:
            basket_slots.append(slot_idx)
    return basket_slots

try:
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1252']
    df = None
    last_error = None
    for encoding in encodings_to_try:
        try:
            df = pd.read_csv('combined.csv', encoding=encoding)
            df = df.replace(r'^\s*$', pd.NA, regex=True)
            df = df.replace('nan', pd.NA)
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            last_error = e
            continue
    if df is None:
        print(f"Error: Unable to read combined.csv. Please check the file format.\nDetails: {str(last_error)}")
        exit()
except Exception as e:
    print(f"Error: Failed to load combined.csv.\nDetails: {str(e)}")
    exit()

if df.empty:
    print("Error: No data found in combined.csv")
    exit()

def is_break_time(slot, semester=None):
    global lunch_breaks
    start, end = slot
    morning_break = (time(10, 30) <= start < time(11, 0))
    lunch_break = False
    if semester:
        base_sem = int(str(semester)[0])
        if base_sem in lunch_breaks:
            lunch_start, lunch_end = lunch_breaks[base_sem]
            lunch_break = (lunch_start <= start < lunch_end)
    else:
        lunch_break = any(lunch_start <= start < lunch_end for lunch_start, lunch_end in lunch_breaks.values())
    return morning_break or lunch_break

def is_lecture_scheduled(timetable, day, start_slot, end_slot):
    for slot in range(start_slot, end_slot):
        if (slot < len(timetable[day]) and timetable[day][slot]['type'] and timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def calculate_required_slots(course):
    l = float(course['L']) if pd.notna(course['L']) else 0
    t = int(course['T']) if pd.notna(course['T']) else 0
    p = int(course['P']) if pd.notna(course['P']) else 0
    s = int(course['S']) if pd.notna(course['S']) else 0
    c = int(course['C']) if pd.notna(course['C']) else 0
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
    lecture_sessions = 0
    if l > 0:
        lecture_sessions = max(1, round(l * 2/3))
    tutorial_sessions = t
    lab_sessions = p // 2
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

def select_faculty(faculty_str):
    if '/' in faculty_str:
        faculty_options = [f.strip() for f in faculty_str.split('/')]
        return faculty_options[0]
    return faculty_str

def check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, course_code=None, activity_type=None):
    component_count = 0
    faculty_courses = set()
    for slot in timetable[day].values():
        if slot['faculty'] == faculty and slot['type'] in ['LEC', 'LAB', 'TUT']:
            slot_code = slot.get('code', '')
            if slot_code:
                if not is_basket_course(slot_code):
                    component_count += 1
                elif slot_code not in faculty_courses:
                    component_count += 1
                    faculty_courses.add(slot_code)
    if course_code and is_basket_course(course_code):
        basket_group = get_basket_group(course_code)
        existing_slots = get_basket_group_slots(timetable, day, basket_group)
        if existing_slots:
            return component_count < 3
    return component_count < 2

def check_faculty_course_gap(professor_schedule, timetable, faculty, course_code, day, start_slot):
    min_gap_hours = 3
    slots_per_hour = 2
    required_gap = min_gap_hours * slots_per_hour
    for i in range(max(0, start_slot - required_gap), start_slot):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    for i in range(start_slot + 1, min(len(TIME_SLOTS), start_slot + required_gap)):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    return True

def load_reserved_slots():
    try:
        reserved_slots_path = os.path.join('tt data', 'reserved_slots.csv')
        if not os.path.exists(reserved_slots_path):
            print("Warning: reserved_slots.csv not found in uploads, no slots will be reserved")
            return {day: {} for day in DAYS}
        df = pd.read_csv(reserved_slots_path)
        reserved = {day: {} for day in DAYS}
        for _, row in df.iterrows():
            day = row['Day']
            start = datetime.strptime(row['Start Time'], '%H:%M').time()
            end = datetime.strptime(row['End Time'], '%H:%M').time()
            department = str(row['Department'])
            semesters = []
            for s in str(row['Semester']).split(';'):
                s = s.strip()
                if s.isdigit():
                    base_sem = int(s)
                    semesters.extend([f"{base_sem}A", f"{base_sem}B", str(base_sem)])
                else:
                    semesters.append(s)
            key = (department, tuple(semesters))
            if day not in reserved:
                reserved[day] = {}
            if key not in reserved[day]:
                reserved[day][key] = []
            reserved[day][key].append((start, end))
        return reserved
    except Exception as e:
        print(f"Warning: Error loading reserved slots: {str(e)}")
        return {day: {} for day in DAYS}

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    if day not in reserved_slots:
        return False
    slot_start, slot_end = slot
    for (dept, semesters), slots in reserved_slots[day].items():
        if dept == 'ALL' or dept == department:
            if str(semester) in semesters or any(str(semester).startswith(s) for s in semesters):
                for reserved_start, reserved_end in slots:
                    if (slot_start >= reserved_start and slot_start < reserved_end) or \
                       (slot_end > reserved_start and slot_end <= reserved_end):
                        return True
    return False

def load_faculty_preferences():
    preferences = {}
    try:
        with open('tt data/FACULTY.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                preferred_days = [d.strip() for d in row['Preferred Days'].split(';')] if row['Preferred Days'] else []
                preferred_times = []
                if row['Preferred Times']:
                    time_ranges = row['Preferred Times'].split(';')
                    for time_range in time_ranges:
                        start, end = time_range.split('-')
                        start_time = datetime.strptime(start.strip(), '%H:%M').time()
                        end_time = datetime.strptime(end.strip(), '%H:%M').time()
                        preferred_times.append((start_time, end_time))
                preferences[row['Name']] = {
                    'preferred_days': preferred_days,
                    'preferred_times': preferred_times
                }
    except FileNotFoundError:
        print("Warning: FACULTY.csv not found, proceeding without faculty preferences")
        return {}
    return preferences

def is_preferred_slot(faculty, day, time_slot, faculty_preferences):
    if faculty not in faculty_preferences:
        return True
    prefs = faculty_preferences[faculty]
    if prefs['preferred_days'] and DAYS[day] not in prefs['preferred_days']:
        return False
    if prefs['preferred_times']:
        slot_start, slot_end = time_slot
        for pref_start, pref_end in prefs['preferred_times']:
            if (slot_start >= pref_start and slot_end <= pref_end):
                return True
        return False
    return True

def get_course_priority(course):
    priority = 0
    code = str(course['Course Code'])
    if pd.notna(course['P']) and course['P'] > 0 and not is_basket_course(code):
        priority += 10
        if 'CS' in code or 'EC' in code:
            priority += 2
    elif is_basket_course(code):
        priority += 1
    elif pd.notna(course['L']) and course['L'] > 2:
        priority += 3
    elif pd.notna(course['T']) and course['T'] > 0:
        priority += 2
    return priority

def get_best_slots(timetable, professor_schedule, faculty, day, duration, reserved_slots, semester, department, faculty_preferences):
    best_slots = []
    preferred_slots = []
    for start_slot in range(len(TIME_SLOTS) - duration + 1):
        slots_free = True
        for i in range(duration):
            current_slot = start_slot + i
            if duration == LAB_DURATION:
                if (current_slot in professor_schedule[faculty][day] or
                    timetable[day][current_slot]['type'] is not None or
                    is_break_time(TIME_SLOTS[current_slot], semester) or
                    is_slot_reserved(TIME_SLOTS[current_slot], DAYS[day], semester, department, reserved_slots)):
                    slots_free = False
                    break
            else:
                if (current_slot in professor_schedule[faculty][day] or
                    (timetable[day][current_slot]['type'] is not None and
                     not is_basket_course(timetable[day][current_slot].get('code', ''))) or
                    is_break_time(TIME_SLOTS[current_slot], semester) or 
                    is_slot_reserved(TIME_SLOTS[current_slot], DAYS[day], semester, department, reserved_slots)):
                    slots_free = False
                    break
        if slots_free:
            if duration == LAB_DURATION:
                slot_time = TIME_SLOTS[start_slot][0]
                if slot_time < time(12, 30):
                    preferred_slots.append(start_slot)
                else:
                    best_slots.append(start_slot)
            else:
                if is_preferred_slot(faculty, day, TIME_SLOTS[start_slot], faculty_preferences):
                    preferred_slots.append(start_slot)
                else:
                    best_slots.append(start_slot)
    return preferred_slots + best_slots

class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty 
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason
    def __eq__(self, other):
        if not isinstance(other, UnscheduledComponent):
            return False
        return (self.department == other.department and
                self.semester == other.semester and
                self.code == other.code and
                self.component_type == other.component_type and
                self.section == other.section)
    def __hash__(self):
        return hash((self.department, self.semester, self.code, self.component_type, self.section))

def unscheduled_reason(course, department, semester, professor_schedule, rooms, component_type, check_attempts):
    faculty = course['Faculty']
    code = str(course['Course Code'])
    faculty_slots_used = 0
    for day in range(len(DAYS)):
        if faculty in professor_schedule and day in professor_schedule[faculty]:
            faculty_slots_used += len(professor_schedule[faculty][day])
    if faculty_slots_used > 20:
        return f"Faculty '{faculty}' already has {faculty_slots_used/2:.1f} hours of teaching scheduled"
    if component_type == 'LAB':
        lab_rooms_available = False
        for _, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                lab_rooms_available = True
                break
        if not lab_rooms_available:
            return "No suitable lab rooms available in the system"
        lab_rooms_free_slots = 0
        for rid, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                total_slots = len(DAYS) * (len(TIME_SLOTS) - LAB_DURATION)
                used_slots = sum(len(room['schedule'].get(day, [])) for day in range(len(DAYS)))
                lab_rooms_free_slots += (total_slots - used_slots)
        if lab_rooms_free_slots < 5:
            return f"Lab rooms almost fully booked ({lab_rooms_free_slots} slots left)"
    if 'total_students' in course and pd.notna(course['total_students']):
        try:
            total_students = int(course['total_students'])
            if total_students > 100:
                large_rooms_available = False
                for _, room in rooms.items():
                    if room['type'].upper() == 'SEATER_120' or room['type'].upper() == 'SEATER_240':
                        large_rooms_available = True
                        break
                if not large_rooms_available:
                    return f"No rooms available with capacity for {total_students} students"
        except (ValueError, TypeError):
            pass
    if check_attempts > 800:
        return f"No suitable timeslot found after {check_attempts} attempts - heavy scheduling conflicts"
    duration_map = {
        'LEC': f"{LECTURE_DURATION/2} hour",
        'LAB': f"{LAB_DURATION/2} hour",
        'TUT': f"{TUTORIAL_DURATION/2} hour"
    }
    duration_str = duration_map.get(component_type, "")
    return f"Could not find compatible {duration_str} timeslot for {code} {component_type} with faculty {faculty}"

def generate_all_timetables():
    global lunch_breaks
    initialize_time_slots()
    reserved_slots = load_reserved_slots()
    faculty_preferences = load_faculty_preferences()
    workbooks = {}
    professor_schedule = {}
    rooms = load_rooms()
    batch_info = load_batch_data()
    unscheduled_components = set()
    subject_colors = [
        "FFB6C1", "98FB98", "87CEFA", "DDA0DD", "F0E68C", 
        "E6E6FA", "FFDAB9", "B0E0E6", "FFA07A", "D8BFD8",
        "AFEEEE", "F08080", "90EE90", "ADD8E6", "FFB6C1"
    ]
    basket_group_colors = {
        'B1': "FF9999",
        'B2': "99FF99",
        'B3': "9999FF",
        'B4': "FFFF99",
        'B5': "FF99FF",
        'B6': "99FFFF",
        'B7': "FFB366",
        'B8': "B366FF",
        'B9': "66FFB3"
    }
    self_study_courses = []
    all_semesters = sorted(set(int(str(sem)[0]) for sem in df['Semester'].unique()))
    lunch_breaks = calculate_lunch_breaks(all_semesters)
    for department in df['Department'].unique():
        wb = Workbook()
        wb.remove(wb.active)
        workbooks[department] = wb
        course_faculty_assignments = {}
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & 
                        (df['Semester'] == semester) &
                        ((df['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                         (df['Schedule'].isna()))].copy()
            if courses.empty:
                continue
            lab_courses = courses[courses['P'] > 0].copy()
            lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
            lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses = courses[courses['P'] == 0].copy()
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)
            courses = pd.concat([lab_courses, non_lab_courses])
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1
            for _, course in courses.iterrows():
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': str(course['Faculty']),
                        'department': department,
                        'semester': semester
                    })
            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                subject_color_map = {}
                course_faculty_map = {}
                color_idx = 0
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            subject_color_map[code] = basket_group_colors.get(basket_group, subject_colors[color_idx % len(subject_colors)])
                        else:
                            subject_color_map[code] = subject_colors[color_idx % len(subject_colors)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1
                courses['priority'] = courses.apply(get_course_priority, axis=1)
                courses = courses.sort_values('priority', ascending=False)
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    if not any(code.startswith(f'B{i}') for i in range(1, 10)):
                        if code in course_faculty_assignments:
                            if '/' in faculty:
                                faculty_options = [f.strip() for f in faculty.split('/')] 
                                available_faculty = [f for f in faculty_options if f not in course_faculty_assignments[code]]
                                if available_faculty:
                                    faculty = available_faculty[0]
                                else:
                                    faculty = select_faculty(faculty)
                        else:
                            faculty = select_faculty(faculty)
                            course_faculty_assignments[code] = [faculty]
                    else:
                        faculty = select_faculty(faculty)
                    lecture_sessions, tutorial_sessions, lab_sessions, _ = calculate_required_slots(course)
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                    for _ in range(lecture_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                attempts += 1
                                continue
                            slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                DAYS[day],
                                                                semester,
                                                                department,
                                                                reserved_slots) 
                                               for i in range(LECTURE_DURATION))
                            if slots_reserved:
                                attempts += 1
                                continue
                            if not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                               department, semester, section, timetable,
                                                               code, 'LEC'):
                                attempts += 1
                                continue
                            slots_free = True
                            for i in range(LECTURE_DURATION):
                                current_slot = start_slot + i
                                if (current_slot in professor_schedule[faculty][day] or 
                                    timetable[day][current_slot]['type'] is not None or
                                    is_break_time(TIME_SLOTS[current_slot], semester)):
                                    slots_free = False
                                    break
                                if current_slot > 0:
                                    if is_lecture_scheduled(timetable, day, 
                                                         max(0, current_slot - BREAK_DURATION), 
                                                         current_slot):
                                        slots_free = False
                                        break
                                if current_slot < len(TIME_SLOTS) - 1:
                                    if is_lecture_scheduled(timetable, day,
                                                         current_slot + 1,
                                                         min(len(TIME_SLOTS), 
                                                             current_slot + BREAK_DURATION + 1)):
                                        slots_free = False
                                        break
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, LECTURE_DURATION, 
                                                          rooms, batch_info, timetable, code)
                                if room_id:
                                    classroom = room_id
                                    for i in range(LECTURE_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'LEC'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            unscheduled_components.add(
                                UnscheduledComponent(department, semester, code, name, 
                                                   faculty, 'LEC', 1, section)
                            )
                    for _ in range(tutorial_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                attempts += 1
                                continue
                            if not check_faculty_daily_components(professor_schedule, faculty, day,
                                                               department, semester, section, timetable,
                                                               code, 'TUT'):
                                attempts += 1
                                continue
                            start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                            slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                DAYS[day],
                                                                semester,
                                                                department,
                                                                reserved_slots) 
                                               for i in range(TUTORIAL_DURATION))
                            if slots_reserved:
                                attempts += 1
                                continue
                            slots_free = True
                            for i in range(TUTORIAL_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    timetable[day][start_slot+i]['type'] is not None or
                                    is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                    slots_free = False
                                    break
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, TUTORIAL_DURATION, 
                                                          rooms, batch_info, timetable, code)
                                if room_id:
                                    classroom = room_id
                                    for i in range(TUTORIAL_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            unscheduled_components.add(
                                UnscheduledComponent(department, semester, code, name,
                                                   faculty, 'TUT', 1, section)
                            )
                    if lab_sessions > 0:
                        room_type = get_required_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            scheduling_reason = ""
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            for day in days:
                                possible_slots = get_best_slots(timetable, professor_schedule, 
                                                              faculty, day, LAB_DURATION, 
                                                              reserved_slots, semester, department, faculty_preferences)
                                for start_slot in possible_slots:
                                    room_id = find_suitable_room(room_type, department, semester,
                                                               day, start_slot, LAB_DURATION,
                                                               rooms, batch_info, timetable, code)
                                    if room_id:
                                        classroom = room_id if ',' not in str(room_id) else f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                                        for i in range(LAB_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                if scheduled:
                                    break
                            if not scheduled:
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name,
                                                       faculty, 'LAB', 1, section,
                                                       "Could not find suitable room and time slot combination")
                                )
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    _, _, _, self_study_sessions = calculate_required_slots(course)
                    if self_study_sessions > 0:
                        if faculty not in professor_schedule:
                            professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                        for _ in range(self_study_sessions):
                            scheduled = False
                            attempts = 0
                            while not scheduled and attempts < 1000:
                                day = random.randint(0, len(DAYS)-1)
                                start_slot = random.randint(0, len(TIME_SLOTS)-SELF_STUDY_DURATION)
                                slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                    DAYS[day],
                                                                    semester,
                                                                    department,
                                                                    reserved_slots) 
                                                   for i in range(SELF_STUDY_DURATION))
                                if slots_reserved:
                                    attempts += 1
                                    continue
                                slots_free = True
                                for i in range(SELF_STUDY_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        timetable[day][start_slot+i]['type'] is not None or
                                        is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                        slots_free = False
                                        break
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, SELF_STUDY_DURATION, 
                                                              rooms, batch_info, timetable, code)
                                    if room_id:
                                        classroom = room_id
                                        for i in range(SELF_STUDY_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'SS'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                attempts += 1
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                ss_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    merge_ranges = []
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        if is_break_time(TIME_SLOTS[slot_idx], semester):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            if code:
                                duration = {
                                    'LEC': LECTURE_DURATION,
                                    'LAB': LAB_DURATION,
                                    'TUT': TUTORIAL_DURATION,
                                    'SS': SELF_STUDY_DURATION
                                }.get(activity_type, 1)
                                if code in subject_color_map:
                                    cell_fill = PatternFill(start_color=subject_color_map[code],
                                                          end_color=subject_color_map[code],
                                                          fill_type="solid")
                                else:
                                    cell_fill = {
                                        'LAB': lab_fill,
                                        'TUT': tut_fill,
                                        'SS': ss_fill,
                                        'LEC': lec_fill
                                    }.get(activity_type, lec_fill)
                                if code and is_basket_course(code):
                                    basket_group = get_basket_group(code)
                                    basket_codes = set()
                                    basket_details = {}
                                    for slot_id, slot_data in timetable[day_idx].items():
                                        slot_code = slot_data.get('code', '')
                                        if (slot_data.get('type') == activity_type and 
                                            get_basket_group(slot_code) == basket_group):
                                            basket_codes.add(slot_code)
                                            if slot_code not in basket_details:
                                                basket_details[slot_code] = {
                                                    'faculty': slot_data['faculty'],
                                                    'room': slot_data['classroom']
                                                }
                                    if basket_codes:
                                        basket_header = f"{basket_group} Courses\n"
                                        codes_str = ', '.join(sorted(basket_codes))
                                        course_details = [
                                            f"{code}: {details['faculty']} ({details['room']})"
                                            for code, details in sorted(basket_details.items())
                                            if code and details['faculty'] and details['room']
                                        ]
                                        cell_value = f"{basket_header}{codes_str}\n" + "\n".join(course_details)
                                else:
                                    cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}"
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40
                current_row = len(DAYS) + 4
                if self_study_courses:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1
                    for course in self_study_courses:
                        if course['department'] == department and course['semester'] == semester:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                    current_row += 2
                dept_unscheduled = [c for c in unscheduled_components 
                                    if c.department == department and 
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]
                if dept_unscheduled:
                    current_row += 2
                    unsch_title = ws.cell(row=current_row, column=1, value="Unscheduled Components")
                    unsch_title.font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2
                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    current_row += 1
                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason or "Could not find suitable slot", None)
                        ]
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
                    current_row += 2
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                legend_headers = ['Subject Code', 'Subject Name', 'Faculty', 'Color']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                for code, color in subject_color_map.items():
                    if code in course_faculty_map:
                        ws.row_dimensions[current_row].height = 25
                        cells = [
                            (code, None),
                            (course_faculty_map[code]['name'], None),
                            (course_faculty_map[code]['faculty'], None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid"))
                        ]
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
    for department, wb in workbooks.items():
        filename = f"timetable_{department}.xlsx"
        wb.save(filename)
        print(f"Timetable for {department} saved as {filename}")
    return [f"timetable_{dept}.xlsx" for dept in workbooks.keys()]

if __name__ == "__main__":
    generate_all_timetables()
    
